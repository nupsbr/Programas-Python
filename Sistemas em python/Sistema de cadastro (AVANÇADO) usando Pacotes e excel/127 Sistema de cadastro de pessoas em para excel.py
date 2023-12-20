from Pacote_uteis import c, cadastro as cd
from openpyxl import Workbook, load_workbook
from time import sleep

# Carregue o arquivo Excel existente (se houver)
try:
    arquivo_excel = load_workbook("Cadastro.xlsx")
    planilha = arquivo_excel.active
except FileNotFoundError:
    # Se o arquivo não existir, crie um novo
    arquivo_excel = Workbook()
    planilha = arquivo_excel.active
    planilha.append(["Nome", "Idade"])

while True:
    try:
        cd.menuCP()
        sleep(1)
        n = int(input(f'{c.br}Sua Opção: {c.am}'))
    except (ValueError, TypeError):
        print(f'{c.vm}ERRO: Por favor, digite um numero inteiro valido.{c.lp}')
        sleep(1)
    else:
        if n == 1:
            c.liam(50)
            # PROGRAMA OPÇÃO 1
            sleep(1)
            print(f'\t{c.az}Carregando planilha...')
            sleep(2)
            print(f'{c.vr}\tPlanilha carregada com sucesso!')
            sleep(1)
            for linha in planilha.iter_rows(min_row=2, values_only=True):
                nome, idade = linha
                print(
                    f'{c.brs}Nome:{c.az}{nome:^35} {c.br}Idade: {c.az}{idade:^5}{c.lp}')
                sleep(0.3)
            sleep(3)
            c.liam(50)
            cd.continuarMN()
        elif n == 2:
            c.liam(50)
            # PROGRAMA OPÇÃO 2
            while True:
                try:
                    nome = str(input('Nome: '))
                    idade = int(input('Idade: '))
                except:
                    print(
                        f'{c.vm}ERRO: Por favor, digite as informações validas.{c.lp}')
                else:
                    planilha.append([nome, idade])
                    sleep(1)
                    print(f'{c.az}\tAdicionando usuário na planilha...')
                    sleep(3)
                    print(
                        f'{c.vr}\tUsuario adicionado com sucesso!!\n\t{c.az}      RETORNANDO AO MENU.')
                    sleep(2)
                    break
        elif n == 3:
            sleep(1)
            # PROGRAMA OPÇÃO 3
            c.liam(50)
            print(f'{c.az}\tEncerrando o Programa...')
            sleep(2)
            print(f'{c.vr}  Programa encerrado com sucesso, Volte sempre!!{c.lp}')
            c.liam(50)
            break
        else:
            c.liam(50)
            print(f'{c.vm}ESSA OPÇÃO NAO EXISTE, TENTE NOVAMENTE!{c.lp}')
            c.liam(50)

        arquivo_excel.save("Cadastro.xlsx")
